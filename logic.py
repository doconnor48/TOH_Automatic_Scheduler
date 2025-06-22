import csv
import random
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.utils import get_column_letter
import io
from openpyxl import load_workbook




class worker:
    def __init__(self, name, rank, offdays, sibling=False):
        self.name = name
        self.rank = rank
        self.offdays = offdays
        self.sibling = sibling
class sibl:
    def __init__(self,name,beach):
        self.name = name
        self.beach = beach


def generate_excel_from_csv(file, previous_schedule_file = None) -> io.BytesIO:
    persons = []
    off_idx1 = None
    off_idx2 = None
    reader = csv.reader(io.StringIO(file.read().decode('utf-8')))
    headers = next(reader)
    name_idx = rank_idx = sibling_idx = None
    no_sibling = False

    for i, header in enumerate(headers):
        header = header.lower()
        if 'name' in header:
            name_idx = i
        elif 'off' in header:
            if off_idx1 is None:
                off_idx1 = i
            else:
                off_idx2 = i
        elif 'rank' in header:
            rank_idx = i
        elif 'sibling' in header:
            sibling_idx = i

    for row in reader:
        name = str(row[name_idx]).strip()
        off_day = [item.strip().lower() for item in str(row[off_idx1]).split(',')]
        if off_idx2 is not None:
            tmp = [item.strip().lower() for item in str(row[off_idx2]).split(',')]
            off_day.extend(tmp)
        rank = str(row[rank_idx]).strip().lower()
        if sibling_idx is not None:
            sibling = row[sibling_idx].strip().lower() == 'yes'
            persons.append(worker(name, rank, off_day, sibling))
        else:
            persons.append(worker(name,rank, off_day, False))

    # Combine and shuffle
    random.shuffle(persons)

    # seperate by ranks
    slts = []
    lts = []
    sgs = []
    rookies = []
    guards = []
    senior_lt_pseuds = ["slt", "senior lieutenant", "senior leuitenant", 'senior liuetenant', 'senor lieutenant', 'senor leuitenant', 'senor liuetenant', 'sen lt', 'senior lt', 'senior lieut','senior liet', 'senor lt', 'senor lieut', 'sen lt', 'slt.', 'sen lt.', 'senior lt.', 'sen lt.', 'senior liet.', 'senor lt.', 'sen lt.']
    lt_pseuds = ['lt', 'lieutenant', 'leuitenant', 'liuetenant', 'liet', 'lieut', 'lut', 'lt.', 'liet.']
    sg_pseuds = ['senior guard', 'sg', 'senor gard', 'senor guard', 'sg.', 'sr guard', 'sr. guard', 'sr gard', 'sr. gard','srg', 'sg.', 'srg.']
    rookie_pseuds = ['rookie', 'rooky', 'rook', 'first year', '1rst year', '1. year', ' 1 year', '1rst. year']
    for person in persons:
        if person.rank in senior_lt_pseuds:
            slts.append(person)
        elif person.rank in lt_pseuds:
            lts.append(person)
        elif person.rank in sg_pseuds:
            sgs.append(person)
        elif person.rank in rookie_pseuds:
            rookies.append(person)
        else:
            guards.append(person)

    # get the number of people needed at each beach
    #These are weights of each beach. Example: Civic, 0.05 means civic gets 5 percent of the entire crew. These are able to be changed, but the total should add up to 1. I just went with my best guess as to how they were distributed
    beaches = [('Civic',0.05), ('Middle',0.035), ('2Chair',0.04), ('Main',0.1), ('7 Chair',0.035), ('Malibu',0.07), ('Nassau 1',0.03), ('Nassau 2',0.08), ('Nassau 5', 0.05), ('Reef', 0.1), ('Anchor', 0.04),('East Lido', 0.035), ('Main Lido', 0.06), ('West Lido', 0.035), ('Lido West', 0.1), ('Surfing Bay', 0.035), ('EAB', 0.07), ('Sea Glades', 0.035)]

    num_people = []
    count = 0
    for i in range(len(beaches)):
        number = np.round(len(persons)*beaches[i][1])
        count += number
        num_people.append(number)

    if count < len(persons):
        num_people[random.randint(0,len(num_people)-1)] += len(persons)- count


    slt_beaches = ['Lido West', 'Reef', 'Main', 'Nassau 2']
    lt_beaches = ['Lido West','Reef', 'Main', 'Malibu', 'Nassau 2', 'EAB', 'Main Lido', 'Civic']
    sg_beaches = ['Sea Glades','Middle', '2Chair', '7 Chair', 'Nassau 1', 'Nassau 5', 'Anchor',
              'Surfing Bay', 'West Lido', 'East Lido']
    rookie_beaches = ['Main', 'Nassau 2', 'Reef', 'Lido West', 'Malibu']
    family_names = {'murphy', 'walter', 'walsh', 'dorn', 'cody', 'rinn', 'pongratz', 'newby', 'russo', 'gutman', 'trzcinski', 'baller', 'favata', 'fitzpatrick', 'canty', 'boccio', 'creagh'}


    def extract_previous_data(previous_schedule_file):
        wb = load_workbook(previous_schedule_file)
        ws = wb.active
        person_to_beach = {}
        for col in ws.iter_cols(min_col = 1):
            curr_beach = None
            for row_idx, cell in enumerate(col):
                if cell.font and cell.font.bold:
                    curr_beach = str(cell.value).strip()
                elif curr_beach and cell.value:
                    name = str(cell.value).strip()
                    if name:
                        person_to_beach[name] = curr_beach
        return person_to_beach
    
                

    beach_names = [b[0] for b in beaches]
    beach_capacity = {beach: int(n) for beach, n in zip(beach_names, num_people)}
    assigned_counts = defaultdict(int)
    off_day_dict = defaultdict(set)
    assignments = defaultdict(list)

    #for slts/lts and sgs
    def count_offday_overlap(beach, person):
        return len(off_day_dict[beach].intersection(person.offdays))

    def assign_rank_group(rank_group, preferred_beaches, beach_quota, already_assigned_families, family_names, beach_names, previous_beach_dic = None):
        for person in rank_group:
            assigned = False
            assigned_sib = False
            preferred_beaches_1 = None
            last_name = person.name.split(' ')
            last_name = last_name[-1].lower()
            allowed_beaches = None
            if last_name in family_names:
                assigned_sib = True
                for sib in already_assigned_families:
                    if sib.name == last_name:
                        beach = sib.beach
                        assigned_sib = False
                        break
                if not assigned_sib:
                    if beach == 'EAB' or beach == 'Sea Glades':
                        preferred_beaches_1 = list(set(preferred_beaches) & set(['EAB', 'Sea Glades']))
                    else:
                        for i in range(len(beach_names)):  
                            if beach_names[i] == beach:
                                ind_low = i-2
                                ind_high = i+2
                                if ind_high > len(beach_names) -1:
                                    allowed_beaches = beach_names[ind_low:len(beach_names)]
                                    break
                                elif ind_low < 0:
                                    allowed_beaches = beach_names[:ind_high+1]
                                    break
                                else:
                                    allowed_beaches = beach_names[ind_low:ind_high+1]
                                    break
                        if allowed_beaches:
                            preferred_beaches_1 = list(set(allowed_beaches) & set(preferred_beaches))
            if preferred_beaches_1 is not None:
                if previous_beach_dic:
                    parts = person.name.split(' ')
                    not_good_beach = previous_beach_dic.get(str(parts[-1]).upper())
                    key =str(parts[-1]).upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper()
                    if key:
                        previous_beach_dic.pop(key)
                    if not_good_beach in preferred_beaches_1:
                        preferred_beaches_1.remove(not_good_beach)
                sorted_beaches = sorted(
                preferred_beaches_1,
                key=lambda b: (
                    count_offday_overlap(b, person),
                    assigned_counts[b] / beach_capacity[b] if beach_capacity[b] > 0 else 1
                ))
            else:
                if previous_beach_dic:
                    parts = person.name.split(' ')
                    not_good_beach = previous_beach_dic.get(str(parts[-1]).upper())
                    key =str(parts[-1]).upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper()
                    if key:
                        previous_beach_dic.pop(key)
                    if not_good_beach in preferred_beaches:
                        preferred_beaches.remove(not_good_beach)
                sorted_beaches = sorted(
                preferred_beaches,
                key=lambda b: (
                    count_offday_overlap(b, person),
                    assigned_counts[b] / beach_capacity[b] if beach_capacity[b] > 0 else 1
                ))

            for beach in sorted_beaches:
                if assigned_counts[beach] < beach_capacity[beach] and assigned_counts[beach] < beach_quota[beach]:
                    assignments[beach].append(person)
                    assigned_counts[beach]+=1
                    off_day_dict[beach].update(person.offdays)
                    assigned = True
                    if assigned_sib:
                        tmp = sibl(last_name,beach)
                        already_assigned_families.append(tmp)
                    break
            if assigned == False:
                for beach in sorted_beaches:
                    if assigned_counts[beach] < beach_capacity[beach]:
                        assignments[beach].append(person)
                        assigned_counts[beach] +=1
                        off_day_dict[beach].update(person.offdays)
                        assigned = True
                        if assigned_sib:
                            tmp = sibl(last_name,beach)
                            already_assigned_families.append(tmp)
                        break

    def assign_rank_group2(rank_group, preferred_beaches, already_assigned_families, family_names, beach_names):
        for person in rank_group:
            assigned = False
            assigned_sib = False
            last_name = person.name.split(' ')
            preferred_beaches_1 = None
            last_name = last_name[-1].lower()
            allowed_beaches = None
            if last_name in family_names:
                assigned_sib = True
                for sib in already_assigned_families:
                    if sib.name == last_name:
                        beach = sib.beach
                        assigned_sib = False
                        break
                if not assigned_sib:
                    if beach == 'EAB' or beach == 'Sea Glades':
                        preferred_beaches_1 = list(set(preferred_beaches) & set(['EAB', 'Sea Glades']))
                    else:
                        for i in range(len(beach_names)):  
                            if beach_names[i] == beach:
                                ind_low = i-2
                                ind_high = i+2
                                if ind_high > len(beach_names) -1:
                                    allowed_beaches = beach_names[ind_low:len(beach_names)]
                                    break
                                elif ind_low < 0:
                                    allowed_beaches = beach_names[:ind_high+1]
                                    break
                                else:
                                    allowed_beaches = beach_names[ind_low:ind_high+1]
                                    break
                        if allowed_beaches:
                            preferred_beaches_1 = list(set(allowed_beaches) & set(preferred_beaches))
            if preferred_beaches_1 is not None:
                if previous_beach_dic:
                    parts = person.name.split(' ')
                    not_good_beach = previous_beach_dic.get(str(parts[-1]).upper())
                    key =str(parts[-1]).upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper()
                    if key:
                        previous_beach_dic.pop(key)
                    if not_good_beach in preferred_beaches_1:
                        preferred_beaches_1.remove(not_good_beach)
                sorted_beaches = sorted(
                preferred_beaches_1,
                key=lambda b: (
                    count_offday_overlap(b, person),
                    assigned_counts[b] / beach_capacity[b] if beach_capacity[b] > 0 else 1
                ))
            else:
                if previous_beach_dic:
                    parts = person.name.split(' ')
                    not_good_beach = previous_beach_dic.get(str(parts[-1]).upper())
                    key =str(parts[-1]).upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper()
                    if not_good_beach:
                        not_good_beach = previous_beach_dic.get(str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper())
                        key =str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper()
                    if key:
                        previous_beach_dic.pop(key)
                    if not_good_beach in preferred_beaches:
                        preferred_beaches.remove(not_good_beach)
                sorted_beaches = sorted(
                preferred_beaches,
                key=lambda b: (
                    count_offday_overlap(b, person),
                    assigned_counts[b] / beach_capacity[b] if beach_capacity[b] > 0 else 1
                ))

            for beach in sorted_beaches:
                if assigned_counts[beach] < beach_capacity[beach]:
                    assignments[beach].append(person)
                    assigned_counts[beach]+=1
                    off_day_dict[beach].update(person.offdays)
                    assigned = True
                    if assigned_sib:
                        tmp = sibl(last_name, beach)
                        already_assigned_families.append(tmp)
                    break
    beach_quota_lt = {
        'Lido West': 3,
        'Reef' : 3,
        'Main' : 3,
        'Malibu' : 3,
        'Nassau 2': 3,
        'EAB' : 2,
        'Main Lido' : 2,
        'Civic': 2
    }
    beach_quota_sg = {
        'Nassau 5': 2,
        'Middle' : 2,
        '2Chair' : 2,
        '7Chair': 1,
        'Nassau 1': 1,
        'Anchor' : 1,
        'Surfing Bay':1,
        'West Lido' : 1,
        'East Lido' : 1,
        'Sea Glades': 1,
        'EAB': 1
    }
    num_of_rookies_per = round(len(rookies) / len(rookie_beaches))
    if num_of_rookies_per * len(rookie_beaches) < len(rookies):
        lido_west_rookies = num_of_rookies_per + len(rookies) - (num_of_rookies_per*len(rookie_beaches))
    else:
        lido_west_rookies = num_of_rookies_per

    beach_quota_rookie = {
        'Main': num_of_rookies_per,
        'Malibu': num_of_rookies_per,
        'Nassau 2': num_of_rookies_per,
        'Reef': num_of_rookies_per,
        'Lido West': lido_west_rookies
    }
    already_assigned_families = []
    random.shuffle(slt_beaches)
    #if sg beaches are to deterministic (not enough sgs)
    #random.shuffle(sg_beaches)

    #generate dictionary from previous excel file, if necessary
    previous_beach_dic = None
    if previous_schedule_file:
        previous_beach_dic = extract_previous_data(previous_schedule_file)
    
    assign_rank_group(slts, slt_beaches, beach_quota_lt, already_assigned_families, family_names, beach_names)
    assign_rank_group(lts, lt_beaches, beach_quota_lt, already_assigned_families, family_names, beach_names)
    assign_rank_group(sgs, sg_beaches, beach_quota_sg, already_assigned_families, family_names, beach_names)
    assign_rank_group2(guards, beach_names, already_assigned_families, family_names, beach_names)
    assign_rank_group(rookies, rookie_beaches, beach_quota_rookie, already_assigned_families, family_names, beach_names)

    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"

    # Create color fills
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # SLT, LT
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")    # SG
    bold_font = Font(bold=True)
    italicized_font = Font(italic=True)

    double_letter_names = ["baller", "favata"]



    half = len(beach_names) // 2  # 9 if you have 18 beaches
    max_row_top = 1  # to track row depth in top half
    max_row_bottom = 1  # to track row depth in bottom half

    for i, beach in enumerate(beach_names):
        people = assignments[beach]
    
        # Determine section: top or bottom half
        if i < half:
            col = i + 1
            row_offset = 1
            max_row_ref = max_row_top
        else:
            col = (i - half) + 1
            row_offset = max_row_top + 3  # leave 2 empty rows between top and bottom
            max_row_ref = max_row_bottom
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 25
        # Write header
        cell = ws.cell(row=row_offset, column=col)
        cell.value = beach
        cell.font = bold_font

        # Write names
        for j, person in enumerate(people, start=1):
            name_cell = ws.cell(row=row_offset + j, column=col)
            parts = person.name.split(' ')
            abbreviated = [day[:3].lower() + "." for day in person.offdays]
            tmp = ', '.join(abbreviated)
            if person.sibling or (parts[-1] in family_names):
                if str(parts[-1]).lower() in double_letter_names:
                    s = str(f"{parts[0][0]}{parts[0][1]}. {parts[-1]} ({tmp})").upper()
                else:
                    s = str(f"{parts[0][0]}. {parts[-1]} ({tmp})").upper()
                name_cell.value = s
            else:
                name_cell.value = str(f"{parts[-1]} ({tmp})").upper()
        
            if person.rank in ("senior lieutenant", "lieutenant"):
                name_cell.fill = yellow_fill
            elif person.rank == "senior guard":
                name_cell.fill = blue_fill
            elif person.rank == 'rookie':
                name_cell.font = italicized_font

        # Update max row trackers
        if i < half:
            max_row_top = max(max_row_top, row_offset + len(people))
        else:
            max_row_bottom = max(max_row_bottom, row_offset + len(people))


    # Determine the column where the key should start (2 columns after the last beach column)
    key_col = (half if len(beach_names) % 2 == 0 else half + 1) + 2  # base it on the wider half
    key_col += 2  # leave 1 column gap from the last beach column
    key_col_letter = get_column_letter(key_col)
    ws.column_dimensions[key_col_letter].width = 25

    # Add the key title
    key_title_cell = ws.cell(row=row_offset, column=key_col)
    key_title_cell.value = "Key"
    key_title_cell.font = bold_font

    # Add the rank descriptions with corresponding colors
    lt_cell = ws.cell(row=row_offset+1, column=key_col)
    lt_cell.value = "Lieutenant"
    lt_cell.fill = yellow_fill

    sg_cell = ws.cell(row=row_offset+2, column=key_col)
    sg_cell.value = "Senior Guard"
    sg_cell.fill = blue_fill

    # Save the workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output